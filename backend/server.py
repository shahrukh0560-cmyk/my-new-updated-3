from fastapi import FastAPI, APIRouter, HTTPException, Depends, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import uuid
import csv
import io
import jwt
import bcrypt
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Literal, Dict, Any
from datetime import datetime, timezone, timedelta

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]
JWT_SECRET = os.environ["JWT_SECRET"]
JWT_ALGORITHM = os.environ.get("JWT_ALGORITHM", "HS256")
JWT_EXPIRE_HOURS = int(os.environ.get("JWT_EXPIRE_HOURS", "168"))
ADMIN_EMAIL = os.environ["ADMIN_EMAIL"]
ADMIN_PASSWORD = os.environ["ADMIN_PASSWORD"]
EMERGENT_LLM_KEY = os.environ.get("EMERGENT_LLM_KEY", "")
DEFAULT_GST_RATE = float(os.environ.get("DEFAULT_GST_RATE", "12"))  # eyewear in India = 12%

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

app = FastAPI(title="OptiCRM API")
api = APIRouter(prefix="/api")
security = HTTPBearer()


# ---------- Helpers ----------
def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()


def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False


def create_token(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "exp": now_utc() + timedelta(hours=JWT_EXPIRE_HOURS),
        "iat": now_utc(),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


async def get_current_user(creds: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    try:
        payload = jwt.decode(creds.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user


def require_roles(*roles: str):
    async def dep(user: dict = Depends(get_current_user)):
        if user.get("role") not in roles:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return user
    return dep


def current_owner_id(user: dict) -> Optional[str]:
    """Returns the tenant (shop owner) id for a user. None for super_admin (sees everything)."""
    if user.get("role") == "super_admin":
        return None
    return user.get("owner_id") or user["id"]


def tenant_filter(user: dict, base: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """Adds tenant scoping to a Mongo query. Super-admin sees everything."""
    q: Dict[str, Any] = dict(base or {})
    oid = current_owner_id(user)
    if oid is not None:
        q["owner_id"] = oid
    return q


def doc(d: dict) -> dict:
    if d and "_id" in d:
        d.pop("_id")
    return d


def next_invoice_no(seq: int) -> str:
    yr = now_utc().strftime("%Y%m")
    return f"INV-{yr}-{seq:05d}"


# ---------- Models ----------
COUNTRY_CURRENCY: Dict[str, Dict[str, str]] = {
    "IN": {"name": "India", "currency": "INR", "symbol": "₹", "locale": "en-IN"},
    "US": {"name": "United States", "currency": "USD", "symbol": "$", "locale": "en-US"},
    "GB": {"name": "United Kingdom", "currency": "GBP", "symbol": "£", "locale": "en-GB"},
    "EU": {"name": "Eurozone", "currency": "EUR", "symbol": "€", "locale": "en-IE"},
    "AE": {"name": "United Arab Emirates", "currency": "AED", "symbol": "د.إ", "locale": "en-AE"},
    "SA": {"name": "Saudi Arabia", "currency": "SAR", "symbol": "﷼", "locale": "en-SA"},
    "AU": {"name": "Australia", "currency": "AUD", "symbol": "A$", "locale": "en-AU"},
    "SG": {"name": "Singapore", "currency": "SGD", "symbol": "S$", "locale": "en-SG"},
    "CA": {"name": "Canada", "currency": "CAD", "symbol": "C$", "locale": "en-CA"},
    "PK": {"name": "Pakistan", "currency": "PKR", "symbol": "₨", "locale": "en-PK"},
    "BD": {"name": "Bangladesh", "currency": "BDT", "symbol": "৳", "locale": "en-BD"},
    "LK": {"name": "Sri Lanka", "currency": "LKR", "symbol": "₨", "locale": "en-LK"},
    "NP": {"name": "Nepal", "currency": "NPR", "symbol": "₨", "locale": "en-NP"},
}


class RegisterIn(BaseModel):
    email: EmailStr
    password: str
    name: str
    country: Optional[str] = "IN"  # ISO-2; auto-maps to currency


class LoginIn(BaseModel):
    email: EmailStr
    password: str


class TokenOut(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: dict


# Branches
class BranchIn(BaseModel):
    name: str
    code: str
    address: Optional[str] = ""
    phone: Optional[str] = ""
    gstin: Optional[str] = ""


# Extended Prescription
RX_TYPES = Literal["distance", "near", "intermediate", "progressive", "bifocal", "contact", "pediatric"]


class PrescriptionIn(BaseModel):
    date: str
    rx_type: RX_TYPES = "distance"
    # OD
    od_sph: Optional[float] = None
    od_cyl: Optional[float] = None
    od_axis: Optional[int] = None
    od_add: Optional[float] = None
    od_prism: Optional[float] = None
    od_base: Optional[str] = None
    od_va: Optional[str] = None
    # OS
    os_sph: Optional[float] = None
    os_cyl: Optional[float] = None
    os_axis: Optional[int] = None
    os_add: Optional[float] = None
    os_prism: Optional[float] = None
    os_base: Optional[str] = None
    os_va: Optional[str] = None
    # Both
    pd: Optional[float] = None
    near_pd: Optional[float] = None
    k_readings: Optional[str] = None
    diagnosis: Optional[str] = None
    doctor_name: Optional[str] = None
    notes: Optional[str] = ""


# Customers
class CustomerIn(BaseModel):
    name: str
    phone: str
    email: Optional[str] = ""
    address: Optional[str] = ""
    dob: Optional[str] = ""
    birthday: Optional[str] = ""  # YYYY-MM-DD (MM-DD used for yearly wish)
    anniversary: Optional[str] = ""  # YYYY-MM-DD
    gstin: Optional[str] = ""
    notes: Optional[str] = ""
    branch_id: Optional[str] = None


# Inventory
class InventoryIn(BaseModel):
    name: str
    category: Literal["frame", "lens", "contact", "accessory"]
    brand: Optional[str] = ""
    model: Optional[str] = ""
    color: Optional[str] = ""
    shape: Optional[str] = ""  # frame: round/square/aviator
    material: Optional[str] = ""
    # Lens-specific
    lens_index: Optional[float] = None  # 1.50, 1.56, 1.60, 1.67, 1.74
    blue_cut: Optional[bool] = False
    photochromic: Optional[bool] = False
    progressive_lens: Optional[bool] = False
    coatings: Optional[str] = ""  # AR, hydrophobic
    # Pricing
    cost: Optional[float] = 0
    price: float
    mrp: Optional[float] = None
    gst_rate: Optional[float] = None  # default 12% for eyewear
    hsn_code: Optional[str] = "9004"  # spectacles
    # Stock
    stock: int = 0
    low_stock_threshold: int = 3
    rack_location: Optional[str] = ""
    supplier: Optional[str] = ""
    warranty_months: Optional[int] = 0
    sku: Optional[str] = ""
    barcode: Optional[str] = ""
    branch_id: Optional[str] = None


class InventoryUpdate(BaseModel):
    name: Optional[str] = None
    brand: Optional[str] = None
    model: Optional[str] = None
    color: Optional[str] = None
    shape: Optional[str] = None
    material: Optional[str] = None
    lens_index: Optional[float] = None
    blue_cut: Optional[bool] = None
    photochromic: Optional[bool] = None
    progressive_lens: Optional[bool] = None
    coatings: Optional[str] = None
    cost: Optional[float] = None
    price: Optional[float] = None
    mrp: Optional[float] = None
    gst_rate: Optional[float] = None
    hsn_code: Optional[str] = None
    stock: Optional[int] = None
    low_stock_threshold: Optional[int] = None
    rack_location: Optional[str] = None
    supplier: Optional[str] = None
    warranty_months: Optional[int] = None
    sku: Optional[str] = None
    barcode: Optional[str] = None


# Orders
ORDER_STATUSES = ["received", "frame_selected", "lens_ordered", "lab_processing", "edging", "fitting", "qc", "ready", "delivered", "cancelled"]


class OrderLineIn(BaseModel):
    item_id: str
    quantity: int = 1


class OrderIn(BaseModel):
    customer_id: str
    prescription_id: Optional[str] = None
    lines: List[OrderLineIn]
    discount: float = 0
    paid: float = 0
    notes: Optional[str] = ""
    branch_id: Optional[str] = None


class OrderStatusIn(BaseModel):
    status: Literal["received", "frame_selected", "lens_ordered", "lab_processing", "edging", "fitting", "qc", "ready", "delivered", "cancelled"]
    note: Optional[str] = ""


class ReminderIn(BaseModel):
    customer_id: str
    channel: Literal["sms", "whatsapp"]
    message: str


# Subscription
SUB_PLANS = {
    "trial": {"id": "trial", "name": "Free Trial", "price": 0, "trial_days": 14, "features": ["1 branch", "100 customers", "Core CRM"]},
    "starter": {"id": "starter", "name": "Starter", "price": 499, "trial_days": 0, "features": ["1 branch", "500 customers", "GST Invoices", "Reports"]},
    "pro": {"id": "pro", "name": "Professional", "price": 1499, "trial_days": 0, "features": ["5 branches", "Unlimited customers", "AI summaries", "WhatsApp marketing"]},
    "enterprise": {"id": "enterprise", "name": "Enterprise", "price": 4999, "trial_days": 0, "features": ["Unlimited branches", "Multi-staff RBAC", "Priority support", "Custom integrations"]},
}


class SubscriptionStartIn(BaseModel):
    plan_id: Literal["trial", "starter", "pro", "enterprise"]


# Super-Admin
class TenantStatusIn(BaseModel):
    status: Literal["active", "suspended"]


class TenantPlanIn(BaseModel):
    plan_id: Literal["trial", "starter", "pro", "enterprise"]
    days: Optional[int] = 30


class AdminBroadcastIn(BaseModel):
    title: str
    message: str
    severity: Literal["info", "warning", "critical"] = "info"


# Staff management (owner/admin creates staff within their tenant)
class StaffCreateIn(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: Literal["admin", "staff"] = "staff"
    branch_id: Optional[str] = None


class StaffUpdateIn(BaseModel):
    name: Optional[str] = None
    role: Optional[Literal["admin", "staff"]] = None
    branch_id: Optional[str] = None
    status: Optional[Literal["active", "suspended"]] = None
    password: Optional[str] = None


# Customer CSV import payload
class CustomerImportRow(BaseModel):
    name: str
    phone: str
    email: Optional[str] = ""
    address: Optional[str] = ""
    dob: Optional[str] = ""
    birthday: Optional[str] = ""
    anniversary: Optional[str] = ""
    gstin: Optional[str] = ""
    notes: Optional[str] = ""
    branch_id: Optional[str] = None


class CustomerImportIn(BaseModel):
    rows: List[CustomerImportRow]
    skip_duplicates: bool = True  # by phone within tenant


# Inventory CSV import payload
class InventoryImportRow(BaseModel):
    name: str
    category: Literal["frame", "lens", "contact", "accessory"]
    price: float
    brand: Optional[str] = ""
    model: Optional[str] = ""
    color: Optional[str] = ""
    shape: Optional[str] = ""
    material: Optional[str] = ""
    lens_index: Optional[float] = None
    blue_cut: Optional[bool] = False
    photochromic: Optional[bool] = False
    progressive_lens: Optional[bool] = False
    coatings: Optional[str] = ""
    cost: Optional[float] = 0
    mrp: Optional[float] = None
    gst_rate: Optional[float] = None
    hsn_code: Optional[str] = "9004"
    stock: int = 0
    low_stock_threshold: int = 3
    rack_location: Optional[str] = ""
    supplier: Optional[str] = ""
    warranty_months: Optional[int] = 0
    sku: Optional[str] = ""
    barcode: Optional[str] = ""
    branch_id: Optional[str] = None


class InventoryImportIn(BaseModel):
    rows: List[InventoryImportRow]
    skip_duplicates: bool = True  # by sku within tenant (when sku present)


# Sales (historical invoices) CSV import payload
class SalesImportRow(BaseModel):
    customer_name: str
    customer_phone: Optional[str] = ""
    total: float
    paid: Optional[float] = 0
    discount: Optional[float] = 0
    gst_amount: Optional[float] = 0
    subtotal: Optional[float] = None
    date: Optional[str] = ""  # ISO date or YYYY-MM-DD
    invoice_no: Optional[str] = ""
    payment_status: Optional[Literal["paid", "partial", "unpaid"]] = None
    notes: Optional[str] = ""
    branch_id: Optional[str] = None


class SalesImportIn(BaseModel):
    rows: List[SalesImportRow]
    skip_duplicates: bool = True  # by invoice_no within tenant


# ---------- Auth ----------
@api.post("/auth/register", response_model=TokenOut)
async def register(body: RegisterIn):
    existing = await db.users.find_one({"email": body.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    uid = str(uuid.uuid4())
    country_code = (body.country or "IN").upper()
    cinfo = COUNTRY_CURRENCY.get(country_code) or COUNTRY_CURRENCY["IN"]
    user = {
        "id": uid,
        "email": body.email.lower(),
        "name": body.name,
        "password_hash": hash_password(body.password),
        "role": "owner",
        "owner_id": uid,  # tenant root
        "status": "active",
        "branch_id": None,
        "country": country_code,
        "currency": cinfo["currency"],
        "currency_symbol": cinfo["symbol"],
        "locale": cinfo["locale"],
        "created_at": now_utc(),
    }
    await db.users.insert_one(user.copy())
    token = create_token(user["id"], user["email"])
    return TokenOut(access_token=token, user={
        "id": user["id"], "email": user["email"], "name": user["name"],
        "role": user["role"], "branch_id": user["branch_id"], "owner_id": user["owner_id"],
        "country": country_code, "currency": cinfo["currency"], "currency_symbol": cinfo["symbol"], "locale": cinfo["locale"],
    })


@api.post("/auth/login", response_model=TokenOut)
async def login(body: LoginIn):
    user = await db.users.find_one({"email": body.email.lower()})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    if user.get("status") == "suspended":
        raise HTTPException(status_code=403, detail="This account has been suspended. Contact OptiCRM support.")
    # Pick currency from the tenant owner (so staff inherit owner's settings)
    oid = user.get("owner_id") or user["id"]
    owner = await db.users.find_one({"id": oid}, {"_id": 0, "country": 1, "currency": 1, "currency_symbol": 1, "locale": 1}) or {}
    token = create_token(user["id"], user["email"])
    return TokenOut(access_token=token, user={
        "id": user["id"], "email": user["email"], "name": user["name"],
        "role": user.get("role", "staff"), "branch_id": user.get("branch_id"),
        "owner_id": oid,
        "country": owner.get("country") or user.get("country") or "IN",
        "currency": owner.get("currency") or user.get("currency") or "INR",
        "currency_symbol": owner.get("currency_symbol") or user.get("currency_symbol") or "₹",
        "locale": owner.get("locale") or user.get("locale") or "en-IN",
    })


@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    # Always return tenant currency/country alongside the user record.
    oid = user.get("owner_id") or user["id"]
    owner = await db.users.find_one({"id": oid}, {"_id": 0, "country": 1, "currency": 1, "currency_symbol": 1, "locale": 1}) or {}
    return {
        **user,
        "country": owner.get("country") or user.get("country") or "IN",
        "currency": owner.get("currency") or user.get("currency") or "INR",
        "currency_symbol": owner.get("currency_symbol") or user.get("currency_symbol") or "₹",
        "locale": owner.get("locale") or user.get("locale") or "en-IN",
    }


@api.get("/countries")
async def list_countries():
    return [{"code": code, **info} for code, info in COUNTRY_CURRENCY.items()]


# ---------- Branches ----------
@api.get("/branches")
async def list_branches(user: dict = Depends(get_current_user)):
    return await db.branches.find(tenant_filter(user), {"_id": 0}).sort("name", 1).to_list(200)


@api.post("/branches")
async def create_branch(body: BranchIn, user: dict = Depends(require_roles("super_admin", "owner", "admin"))):
    b = {"id": str(uuid.uuid4()), **body.model_dump(), "owner_id": current_owner_id(user), "created_at": now_utc()}
    await db.branches.insert_one(b.copy())
    return doc(b)


@api.put("/branches/{bid}")
async def update_branch(bid: str, body: BranchIn, user: dict = Depends(require_roles("super_admin", "owner", "admin"))):
    res = await db.branches.update_one(tenant_filter(user, {"id": bid}), {"$set": body.model_dump()})
    if res.matched_count == 0:
        raise HTTPException(404, "Not found")
    return await db.branches.find_one({"id": bid}, {"_id": 0})


@api.delete("/branches/{bid}")
async def delete_branch(bid: str, user: dict = Depends(require_roles("super_admin", "owner", "admin"))):
    await db.branches.delete_one(tenant_filter(user, {"id": bid}))
    return {"ok": True}


# ---------- Customers ----------
@api.get("/customers")
async def list_customers(q: Optional[str] = None, branch_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    query: Dict[str, Any] = tenant_filter(user)
    if q:
        query["$or"] = [
            {"name": {"$regex": q, "$options": "i"}},
            {"phone": {"$regex": q, "$options": "i"}},
            {"email": {"$regex": q, "$options": "i"}},
        ]
    if branch_id:
        query["branch_id"] = branch_id
    cursor = db.customers.find(query, {"_id": 0}).sort("created_at", -1).limit(1000)
    return await cursor.to_list(1000)


@api.post("/customers")
async def create_customer(body: CustomerIn, user: dict = Depends(get_current_user)):
    c = {
        "id": str(uuid.uuid4()),
        **body.model_dump(),
        "owner_id": current_owner_id(user),
        "prescriptions": [],
        "loyalty_points": 0,
        "created_at": now_utc(),
        "last_visit": now_utc(),
    }
    await db.customers.insert_one(c.copy())
    return doc(c)


@api.get("/customers/{cid}")
async def get_customer(cid: str, user: dict = Depends(get_current_user)):
    c = await db.customers.find_one(tenant_filter(user, {"id": cid}), {"_id": 0})
    if not c:
        raise HTTPException(404, "Customer not found")
    return c


@api.put("/customers/{cid}")
async def update_customer(cid: str, body: CustomerIn, user: dict = Depends(get_current_user)):
    res = await db.customers.update_one(tenant_filter(user, {"id": cid}), {"$set": body.model_dump()})
    if res.matched_count == 0:
        raise HTTPException(404, "Not found")
    return await db.customers.find_one({"id": cid}, {"_id": 0})


@api.delete("/customers/{cid}")
async def delete_customer(cid: str, user: dict = Depends(get_current_user)):
    await db.customers.delete_one(tenant_filter(user, {"id": cid}))
    return {"ok": True}


@api.post("/customers/{cid}/prescriptions")
async def add_prescription(cid: str, body: PrescriptionIn, user: dict = Depends(get_current_user)):
    c = await db.customers.find_one(tenant_filter(user, {"id": cid}))
    if not c:
        raise HTTPException(404, "Customer not found")
    rx = {"id": str(uuid.uuid4()), **body.model_dump(), "ai_summary": "", "created_at": now_utc()}
    await db.customers.update_one(
        {"id": cid},
        {"$push": {"prescriptions": rx}, "$set": {"last_visit": now_utc()}},
    )
    return rx


@api.delete("/customers/{cid}/prescriptions/{rx_id}")
async def delete_prescription(cid: str, rx_id: str, user: dict = Depends(get_current_user)):
    await db.customers.update_one(tenant_filter(user, {"id": cid}), {"$pull": {"prescriptions": {"id": rx_id}}})
    return {"ok": True}


@api.post("/customers/{cid}/prescriptions/{rx_id}/ai-summary")
async def ai_summary(cid: str, rx_id: str, user: dict = Depends(get_current_user)):
    c = await db.customers.find_one({"id": cid}, {"_id": 0})
    if not c:
        raise HTTPException(404, "Customer not found")
    rx = next((r for r in c.get("prescriptions", []) if r["id"] == rx_id), None)
    if not rx:
        raise HTTPException(404, "Prescription not found")

    if not EMERGENT_LLM_KEY:
        return {"ai_summary": "", "error": "AI key not configured"}
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"rx-{rx_id}",
            system_message=(
                "You are an optometry assistant. Given a prescription, write a concise 2-3 sentence "
                "plain-English summary highlighting the vision correction (myopia/hyperopia/astigmatism/presbyopia) "
                "and any clinically notable values. No medical advice."
            ),
        ).with_model("gemini", "gemini-3-flash-preview")
        text = (
            f"Customer: {c.get('name')}\n"
            f"Type: {rx.get('rx_type')}\n"
            f"Date: {rx.get('date')}\n"
            f"OD: SPH {rx.get('od_sph')}, CYL {rx.get('od_cyl')}, AXIS {rx.get('od_axis')}, ADD {rx.get('od_add')}, VA {rx.get('od_va')}\n"
            f"OS: SPH {rx.get('os_sph')}, CYL {rx.get('os_cyl')}, AXIS {rx.get('os_axis')}, ADD {rx.get('os_add')}, VA {rx.get('os_va')}\n"
            f"PD: {rx.get('pd')}\nDiagnosis: {rx.get('diagnosis') or '-'}"
        )
        resp = await chat.send_message(UserMessage(text=text))
        summary = str(resp).strip()
    except Exception as e:
        logging.exception("AI summary failed")
        err = str(e)
        friendly = "AI service unavailable. Top up your Emergent LLM key balance." if "budget" in err.lower() else "AI service unavailable right now."
        return {"ai_summary": "", "error": friendly}

    await db.customers.update_one({"id": cid, "prescriptions.id": rx_id}, {"$set": {"prescriptions.$.ai_summary": summary}})
    return {"ai_summary": summary}


# ---------- Inventory ----------
@api.get("/inventory")
async def list_inventory(
    category: Optional[str] = None,
    low_stock: bool = False,
    branch_id: Optional[str] = None,
    q: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    query: Dict[str, Any] = tenant_filter(user)
    if category:
        query["category"] = category
    if branch_id:
        query["branch_id"] = branch_id
    if q:
        query["$or"] = [{"name": {"$regex": q, "$options": "i"}}, {"brand": {"$regex": q, "$options": "i"}}, {"sku": q}, {"barcode": q}]
    cursor = db.inventory.find(query, {"_id": 0}).sort("name", 1).limit(2000)
    items = await cursor.to_list(2000)
    if low_stock:
        items = [i for i in items if i.get("stock", 0) <= i.get("low_stock_threshold", 3)]
    return items


@api.get("/inventory/barcode/{code}")
async def find_by_barcode(code: str, user: dict = Depends(get_current_user)):
    item = await db.inventory.find_one(tenant_filter(user, {"$or": [{"barcode": code}, {"sku": code}]}), {"_id": 0})
    if not item:
        raise HTTPException(404, "No inventory item with that code")
    return item


@api.post("/inventory")
async def create_item(body: InventoryIn, user: dict = Depends(get_current_user)):
    payload = body.model_dump()
    if payload.get("gst_rate") is None:
        payload["gst_rate"] = DEFAULT_GST_RATE
    item = {"id": str(uuid.uuid4()), **payload, "owner_id": current_owner_id(user), "created_at": now_utc()}
    await db.inventory.insert_one(item.copy())
    return doc(item)


@api.put("/inventory/{iid}")
async def update_item(iid: str, body: InventoryUpdate, user: dict = Depends(get_current_user)):
    update = {k: v for k, v in body.model_dump().items() if v is not None}
    res = await db.inventory.update_one(tenant_filter(user, {"id": iid}), {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(404, "Not found")
    return await db.inventory.find_one({"id": iid}, {"_id": 0})


@api.delete("/inventory/{iid}")
async def delete_item(iid: str, user: dict = Depends(get_current_user)):
    await db.inventory.delete_one(tenant_filter(user, {"id": iid}))
    return {"ok": True}


# ---------- Orders ----------
@api.post("/orders")
async def create_order(body: OrderIn, user: dict = Depends(get_current_user)):
    oid = current_owner_id(user)
    customer = await db.customers.find_one(tenant_filter(user, {"id": body.customer_id}), {"_id": 0})
    if not customer:
        raise HTTPException(404, "Customer not found")

    lines_detail = []
    subtotal = 0.0
    gst_amount = 0.0
    for line in body.lines:
        item = await db.inventory.find_one(tenant_filter(user, {"id": line.item_id}), {"_id": 0})
        if not item:
            raise HTTPException(404, f"Inventory item {line.item_id} not found")
        if item.get("stock", 0) < line.quantity:
            raise HTTPException(400, f"Insufficient stock for {item['name']}")
        line_subtotal = float(item["price"]) * line.quantity
        rate = float(item.get("gst_rate") or DEFAULT_GST_RATE)
        line_gst = line_subtotal * rate / 100.0
        subtotal += line_subtotal
        gst_amount += line_gst
        lines_detail.append({
            "item_id": item["id"],
            "name": item["name"],
            "category": item["category"],
            "hsn_code": item.get("hsn_code"),
            "price": item["price"],
            "quantity": line.quantity,
            "gst_rate": rate,
            "gst_amount": round(line_gst, 2),
            "total": round(line_subtotal + line_gst, 2),
        })

    discount = float(body.discount)
    total = max(0.0, subtotal + gst_amount - discount)
    due = max(0.0, total - float(body.paid))
    status_val = "paid" if due <= 0 and total > 0 else ("partial" if body.paid > 0 else "unpaid")

    # invoice number sequence
    counter = await db.counters.find_one_and_update(
        {"id": "invoice"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,
    )
    seq = (counter or {}).get("seq", 1)
    invoice_no = next_invoice_no(seq)

    timeline = [{"status": "received", "at": now_utc(), "note": "Order placed", "by": user.get("name")}]

    order = {
        "id": str(uuid.uuid4()),
        "invoice_no": invoice_no,
        "owner_id": oid,
        "customer_id": body.customer_id,
        "customer_name": customer.get("name"),
        "customer_phone": customer.get("phone"),
        "customer_gstin": customer.get("gstin"),
        "customer_address": customer.get("address"),
        "branch_id": body.branch_id or customer.get("branch_id"),
        "prescription_id": body.prescription_id,
        "lines": lines_detail,
        "subtotal": round(subtotal, 2),
        "gst_amount": round(gst_amount, 2),
        "discount": round(discount, 2),
        "total": round(total, 2),
        "paid": round(float(body.paid), 2),
        "due": round(due, 2),
        "payment_status": status_val,
        "fulfillment_status": "received",
        "timeline": timeline,
        "notes": body.notes,
        "created_at": now_utc(),
    }
    await db.orders.insert_one(order.copy())
    for line in body.lines:
        await db.inventory.update_one({"id": line.item_id}, {"$inc": {"stock": -line.quantity}})
    # loyalty points: 1 point per ₹100 paid
    pts = int(float(body.paid) // 100)
    if pts:
        await db.customers.update_one({"id": body.customer_id}, {"$inc": {"loyalty_points": pts}, "$set": {"last_visit": now_utc()}})
    else:
        await db.customers.update_one({"id": body.customer_id}, {"$set": {"last_visit": now_utc()}})
    return doc(order)


@api.get("/orders")
async def list_orders(customer_id: Optional[str] = None, branch_id: Optional[str] = None, status: Optional[str] = None, user: dict = Depends(get_current_user)):
    q: Dict[str, Any] = tenant_filter(user)
    if customer_id:
        q["customer_id"] = customer_id
    if branch_id:
        q["branch_id"] = branch_id
    if status:
        q["fulfillment_status"] = status
    cursor = db.orders.find(q, {"_id": 0}).sort("created_at", -1).limit(1000)
    return await cursor.to_list(1000)


@api.get("/orders/{oid}")
async def get_order(oid: str, user: dict = Depends(get_current_user)):
    o = await db.orders.find_one(tenant_filter(user, {"id": oid}), {"_id": 0})
    if not o:
        raise HTTPException(404, "Not found")
    return o


@api.post("/orders/{oid}/payment")
async def add_payment(oid: str, amount: float, user: dict = Depends(get_current_user)):
    o = await db.orders.find_one(tenant_filter(user, {"id": oid}))
    if not o:
        raise HTTPException(404, "Not found")
    new_paid = float(o.get("paid", 0)) + amount
    new_due = max(0.0, float(o["total"]) - new_paid)
    new_status = "paid" if new_due <= 0 else ("partial" if new_paid > 0 else "unpaid")
    await db.orders.update_one({"id": oid}, {"$set": {"paid": round(new_paid, 2), "due": round(new_due, 2), "payment_status": new_status}})
    pts = int(amount // 100)
    if pts:
        await db.customers.update_one({"id": o["customer_id"]}, {"$inc": {"loyalty_points": pts}})
    return await db.orders.find_one({"id": oid}, {"_id": 0})


@api.post("/orders/{oid}/status")
async def update_order_status(oid: str, body: OrderStatusIn, user: dict = Depends(get_current_user)):
    o = await db.orders.find_one(tenant_filter(user, {"id": oid}), {"_id": 0})
    if not o:
        raise HTTPException(404, "Not found")
    event = {"status": body.status, "at": now_utc(), "note": body.note or "", "by": user.get("name")}
    await db.orders.update_one(
        {"id": oid},
        {"$set": {"fulfillment_status": body.status}, "$push": {"timeline": event}},
    )
    return await db.orders.find_one({"id": oid}, {"_id": 0})


# ---------- Reminders (MOCK) ----------
@api.post("/reminders")
async def send_reminder(body: ReminderIn, user: dict = Depends(get_current_user)):
    c = await db.customers.find_one(tenant_filter(user, {"id": body.customer_id}), {"_id": 0})
    if not c:
        raise HTTPException(404, "Customer not found")
    rec = {
        "id": str(uuid.uuid4()),
        "owner_id": current_owner_id(user),
        "customer_id": body.customer_id,
        "customer_name": c.get("name"),
        "channel": body.channel,
        "message": body.message,
        "phone": c.get("phone"),
        "status": "sent_mock",
        "sent_at": now_utc(),
    }
    await db.reminders.insert_one(rec.copy())
    return doc(rec)


@api.get("/reminders")
async def list_reminders(user: dict = Depends(get_current_user)):
    return await db.reminders.find(tenant_filter(user), {"_id": 0}).sort("sent_at", -1).limit(500).to_list(500)


# ---------- Dashboard ----------
@api.get("/dashboard")
async def dashboard(branch_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    today = now_utc().replace(hour=0, minute=0, second=0, microsecond=0)
    month_start = today.replace(day=1)
    base: Dict[str, Any] = tenant_filter(user)
    if branch_id:
        base["branch_id"] = branch_id

    today_q = {**base, "created_at": {"$gte": today}}
    month_q = {**base, "created_at": {"$gte": month_start}}

    orders_today = await db.orders.find(today_q, {"_id": 0}).to_list(2000)
    orders_month = await db.orders.find(month_q, {"_id": 0}).to_list(5000)
    open_orders = await db.orders.find({**base, "payment_status": {"$ne": "paid"}}, {"_id": 0}).to_list(5000)

    revenue_today = sum(o.get("paid", 0) for o in orders_today)
    revenue_month = sum(o.get("paid", 0) for o in orders_month)
    pending_due = sum(o.get("due", 0) for o in open_orders)
    gst_collected_month = sum(o.get("gst_amount", 0) for o in orders_month)

    customers_count = await db.customers.count_documents(base)
    inventory_count = await db.inventory.count_documents(base if branch_id else {})

    low_stock_all = await db.inventory.find(base if branch_id else {}, {"_id": 0}).limit(1000).to_list(1000)
    low_stock = [i for i in low_stock_all if i.get("stock", 0) <= i.get("low_stock_threshold", 3)][:10]

    recent_customers = await db.customers.find(base, {"_id": 0}).sort("last_visit", -1).limit(5).to_list(5)
    recent_orders = await db.orders.find(base, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)

    # Pipeline buckets
    pipeline = {}
    for s in ORDER_STATUSES:
        pipeline[s] = await db.orders.count_documents({**base, "fulfillment_status": s})

    # Repair orders summary
    repair_total = await db.repair_orders.count_documents(base)
    repair_open = await db.repair_orders.count_documents({**base, "status": {"$nin": ["delivered", "cancelled"]}})
    repair_ready = await db.repair_orders.count_documents({**base, "status": "ready"})

    # Birthday / Anniversary today
    today_md = now_utc().strftime("%m-%d")
    cs_all = await db.customers.find(base, {"_id": 0}).to_list(20000)
    birthdays_today = [c for c in cs_all if _md_key(c.get("birthday") or "") == today_md or _md_key(c.get("dob") or "") == today_md][:10]
    anniv_today = [c for c in cs_all if _md_key(c.get("anniversary") or "") == today_md][:10]

    return {
        "revenue_today": revenue_today,
        "revenue_month": revenue_month,
        "pending_due": pending_due,
        "gst_collected_month": round(gst_collected_month, 2),
        "orders_today": len(orders_today),
        "orders_month": len(orders_month),
        "customers_count": customers_count,
        "inventory_count": inventory_count,
        "low_stock": low_stock,
        "recent_customers": recent_customers,
        "recent_orders": recent_orders,
        "pipeline": pipeline,
        "repair_total": repair_total,
        "repair_open": repair_open,
        "repair_ready": repair_ready,
        "birthdays_today": birthdays_today,
        "anniversaries_today": anniv_today,
    }


# ---------- Reports ----------
@api.get("/reports/sales")
async def report_sales(start: Optional[str] = None, end: Optional[str] = None, branch_id: Optional[str] = None, period: Optional[str] = None, user: dict = Depends(get_current_user)):
    q: Dict[str, Any] = tenant_filter(user)
    if branch_id:
        q["branch_id"] = branch_id
    if start or end:
        rng = {}
        if start:
            rng["$gte"] = datetime.fromisoformat(start).replace(tzinfo=timezone.utc)
        if end:
            rng["$lte"] = datetime.fromisoformat(end).replace(tzinfo=timezone.utc)
        q["created_at"] = rng
    orders = await db.orders.find(q, {"_id": 0}).sort("created_at", -1).to_list(5000)
    total_revenue = sum(o.get("paid", 0) for o in orders)
    total_due = sum(o.get("due", 0) for o in orders)
    total_gst = sum(o.get("gst_amount", 0) for o in orders)
    total_discount = sum(o.get("discount", 0) for o in orders)

    # Optional time-series (period = "daily" | "monthly" | "yearly")
    series: List[Dict[str, Any]] = []
    p = (period or "").lower()
    if p in ("daily", "monthly", "yearly"):
        buckets: Dict[str, Dict[str, float]] = {}
        for o in orders:
            dt = o.get("created_at")
            if not isinstance(dt, datetime):
                continue
            if p == "daily":
                key = dt.strftime("%Y-%m-%d")
            elif p == "monthly":
                key = dt.strftime("%Y-%m")
            else:  # yearly
                key = dt.strftime("%Y")
            b = buckets.setdefault(key, {"period": key, "orders": 0, "revenue": 0.0, "due": 0.0, "gst": 0.0, "discount": 0.0, "total": 0.0})
            b["orders"] += 1
            b["revenue"] += float(o.get("paid", 0) or 0)
            b["due"] += float(o.get("due", 0) or 0)
            b["gst"] += float(o.get("gst_amount", 0) or 0)
            b["discount"] += float(o.get("discount", 0) or 0)
            b["total"] += float(o.get("total", 0) or 0)
        series = sorted(buckets.values(), key=lambda r: r["period"])
        for r in series:
            for k in ("revenue", "due", "gst", "discount", "total"):
                r[k] = round(r[k], 2)

    return {
        "total_orders": len(orders),
        "total_revenue": round(total_revenue, 2),
        "total_due": round(total_due, 2),
        "total_gst": round(total_gst, 2),
        "total_discount": round(total_discount, 2),
        "orders": orders[:200],
        "series": series,
        "period": p or None,
    }


@api.get("/reports/gst")
async def report_gst(start: Optional[str] = None, end: Optional[str] = None, user: dict = Depends(get_current_user)):
    q: Dict[str, Any] = tenant_filter(user)
    if start or end:
        rng = {}
        if start:
            rng["$gte"] = datetime.fromisoformat(start).replace(tzinfo=timezone.utc)
        if end:
            rng["$lte"] = datetime.fromisoformat(end).replace(tzinfo=timezone.utc)
        q["created_at"] = rng
    orders = await db.orders.find(q, {"_id": 0}).to_list(5000)
    # bucket by HSN/GST rate
    buckets: Dict[str, Dict[str, float]] = {}
    for o in orders:
        for line in o.get("lines", []):
            key = f"{line.get('hsn_code', '9004')}@{line.get('gst_rate', DEFAULT_GST_RATE)}"
            b = buckets.setdefault(key, {"hsn_code": line.get("hsn_code", "9004"), "gst_rate": line.get("gst_rate", DEFAULT_GST_RATE), "taxable": 0, "gst": 0, "lines": 0})
            line_sub = float(line.get("price", 0)) * int(line.get("quantity", 0))
            b["taxable"] += line_sub
            b["gst"] += float(line.get("gst_amount", 0))
            b["lines"] += 1
    rows = [{"key": k, **v, "taxable": round(v["taxable"], 2), "gst": round(v["gst"], 2)} for k, v in buckets.items()]
    return {
        "rows": rows,
        "total_taxable": round(sum(r["taxable"] for r in rows), 2),
        "total_gst": round(sum(r["gst"] for r in rows), 2),
        "total_orders": len(orders),
    }


@api.get("/reports/inventory")
async def report_inventory(user: dict = Depends(get_current_user)):
    items = await db.inventory.find(tenant_filter(user), {"_id": 0}).to_list(5000)
    by_cat: Dict[str, Dict[str, float]] = {}
    total_value = 0.0
    low = 0
    out_of_stock = 0
    for it in items:
        cat = it.get("category", "other")
        v = float(it.get("cost") or 0) * int(it.get("stock", 0))
        total_value += v
        if it.get("stock", 0) == 0:
            out_of_stock += 1
        elif it.get("stock", 0) <= it.get("low_stock_threshold", 3):
            low += 1
        b = by_cat.setdefault(cat, {"count": 0, "stock": 0, "value": 0})
        b["count"] += 1
        b["stock"] += int(it.get("stock", 0))
        b["value"] += v
    return {
        "total_items": len(items),
        "total_value": round(total_value, 2),
        "low_stock_count": low,
        "out_of_stock_count": out_of_stock,
        "by_category": {k: {**v, "value": round(v["value"], 2)} for k, v in by_cat.items()},
    }


@api.get("/reports/sales.csv")
async def report_sales_csv(start: Optional[str] = None, end: Optional[str] = None, branch_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    q: Dict[str, Any] = tenant_filter(user)
    if branch_id:
        q["branch_id"] = branch_id
    if start or end:
        rng = {}
        if start:
            rng["$gte"] = datetime.fromisoformat(start).replace(tzinfo=timezone.utc)
        if end:
            rng["$lte"] = datetime.fromisoformat(end).replace(tzinfo=timezone.utc)
        q["created_at"] = rng
    orders = await db.orders.find(q, {"_id": 0}).sort("created_at", -1).to_list(5000)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Invoice", "Date", "Customer", "Phone", "Subtotal", "GST", "Discount", "Total", "Paid", "Due", "Payment Status", "Fulfillment"])
    for o in orders:
        w.writerow([
            o.get("invoice_no", o["id"][:8]),
            o["created_at"].strftime("%Y-%m-%d %H:%M") if isinstance(o.get("created_at"), datetime) else o.get("created_at"),
            o.get("customer_name"),
            o.get("customer_phone"),
            o.get("subtotal", 0),
            o.get("gst_amount", 0),
            o.get("discount", 0),
            o.get("total", 0),
            o.get("paid", 0),
            o.get("due", 0),
            o.get("payment_status"),
            o.get("fulfillment_status"),
        ])
    return Response(buf.getvalue(), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=sales.csv"})


# ---------- Subscription (MOCK Razorpay) ----------
@api.get("/subscription/plans")
async def list_plans():
    return list(SUB_PLANS.values())


@api.get("/subscription/me")
async def my_subscription(user: dict = Depends(get_current_user)):
    sub = await db.subscriptions.find_one({"user_id": user["id"]}, {"_id": 0})
    if not sub:
        # default trial
        sub = {
            "id": str(uuid.uuid4()),
            "user_id": user["id"],
            "plan_id": "trial",
            "status": "active",
            "started_at": now_utc(),
            "expires_at": now_utc() + timedelta(days=14),
            "auto_renew": False,
            "billing_history": [],
        }
        await db.subscriptions.insert_one(sub.copy())
        sub.pop("_id", None)
    plan = SUB_PLANS.get(sub["plan_id"])
    return {**sub, "plan": plan}


@api.post("/subscription/start")
async def start_subscription(body: SubscriptionStartIn, user: dict = Depends(get_current_user)):
    plan = SUB_PLANS.get(body.plan_id)
    if not plan:
        raise HTTPException(404, "Unknown plan")
    # MOCKED Razorpay order — no real charge
    mock_payment = {
        "id": f"pay_mock_{uuid.uuid4().hex[:12]}",
        "razorpay_order_id": f"order_mock_{uuid.uuid4().hex[:14]}",
        "amount": plan["price"],
        "currency": "INR",
        "status": "captured_mock",
        "at": now_utc(),
    }
    expires = now_utc() + timedelta(days=30 if plan["price"] > 0 else plan.get("trial_days", 14))
    update = {
        "plan_id": body.plan_id,
        "status": "active",
        "started_at": now_utc(),
        "expires_at": expires,
        "auto_renew": plan["price"] > 0,
    }
    await db.subscriptions.update_one(
        {"user_id": user["id"]},
        {"$set": update, "$push": {"billing_history": mock_payment}, "$setOnInsert": {"id": str(uuid.uuid4()), "user_id": user["id"]}},
        upsert=True,
    )
    sub = await db.subscriptions.find_one({"user_id": user["id"]}, {"_id": 0})
    return {**sub, "plan": plan, "mock_payment": mock_payment, "note": "Razorpay integration is MOCKED. Add RAZORPAY_KEY_ID/SECRET to enable live charges."}


# ---------- Super-Admin (Platform Owner) ----------
def _is_super(user: dict):
    if user.get("role") != "super_admin":
        raise HTTPException(403, "Super-admin only")


async def _tenant_metrics(owner_id: str) -> dict:
    customers = await db.customers.count_documents({"owner_id": owner_id})
    inventory = await db.inventory.count_documents({"owner_id": owner_id})
    orders = await db.orders.find({"owner_id": owner_id}, {"_id": 0, "paid": 1, "total": 1, "due": 1, "gst_amount": 1, "created_at": 1}).to_list(5000)
    revenue = sum(o.get("paid", 0) for o in orders)
    due = sum(o.get("due", 0) for o in orders)
    gst = sum(o.get("gst_amount", 0) for o in orders)
    branches = await db.branches.count_documents({"owner_id": owner_id})
    staff = await db.users.count_documents({"owner_id": owner_id})
    sub = await db.subscriptions.find_one({"user_id": owner_id}, {"_id": 0})
    return {
        "customers": customers,
        "inventory": inventory,
        "orders": len(orders),
        "revenue": round(revenue, 2),
        "due": round(due, 2),
        "gst": round(gst, 2),
        "branches": branches,
        "staff": staff,
        "subscription": sub,
    }


@api.get("/admin/metrics")
async def admin_metrics(user: dict = Depends(get_current_user)):
    _is_super(user)
    total_tenants = await db.users.count_documents({"role": "owner"})
    active_tenants = await db.users.count_documents({"role": "owner", "status": {"$ne": "suspended"}})
    suspended_tenants = await db.users.count_documents({"role": "owner", "status": "suspended"})
    total_customers = await db.customers.count_documents({})
    total_inventory = await db.inventory.count_documents({})
    total_orders = await db.orders.count_documents({})
    orders = await db.orders.find({}, {"_id": 0, "paid": 1, "gst_amount": 1}).to_list(20000)
    gmv = sum(o.get("paid", 0) for o in orders)
    gst = sum(o.get("gst_amount", 0) for o in orders)

    # MRR = active paid subscriptions
    subs = await db.subscriptions.find({"status": "active"}, {"_id": 0}).to_list(5000)
    mrr = 0.0
    plan_breakdown: Dict[str, int] = {p: 0 for p in SUB_PLANS}
    for s in subs:
        pid = s.get("plan_id", "trial")
        plan_breakdown[pid] = plan_breakdown.get(pid, 0) + 1
        mrr += SUB_PLANS.get(pid, {}).get("price", 0)

    # New tenants last 7d & 30d
    week_ago = now_utc() - timedelta(days=7)
    month_ago = now_utc() - timedelta(days=30)
    new_week = await db.users.count_documents({"role": "owner", "created_at": {"$gte": week_ago}})
    new_month = await db.users.count_documents({"role": "owner", "created_at": {"$gte": month_ago}})

    return {
        "total_tenants": total_tenants,
        "active_tenants": active_tenants,
        "suspended_tenants": suspended_tenants,
        "new_tenants_7d": new_week,
        "new_tenants_30d": new_month,
        "total_customers": total_customers,
        "total_inventory": total_inventory,
        "total_orders": total_orders,
        "platform_gmv": round(gmv, 2),
        "platform_gst": round(gst, 2),
        "mrr": round(mrr, 2),
        "arr": round(mrr * 12, 2),
        "plan_breakdown": plan_breakdown,
    }


@api.get("/admin/tenants")
async def admin_list_tenants(q: Optional[str] = None, status: Optional[str] = None, user: dict = Depends(get_current_user)):
    _is_super(user)
    query: Dict[str, Any] = {"role": "owner"}
    if status:
        if status == "suspended":
            query["status"] = "suspended"
        elif status == "active":
            query["status"] = {"$ne": "suspended"}
    if q:
        query["$or"] = [
            {"email": {"$regex": q, "$options": "i"}},
            {"name": {"$regex": q, "$options": "i"}},
        ]
    owners = await db.users.find(query, {"_id": 0, "password_hash": 0}).sort("created_at", -1).limit(500).to_list(500)
    out = []
    for o in owners:
        m = await _tenant_metrics(o["id"])
        out.append({**o, "metrics": {
            "customers": m["customers"],
            "orders": m["orders"],
            "revenue": m["revenue"],
            "branches": m["branches"],
        }, "subscription": m["subscription"]})
    return out


@api.get("/admin/tenants/{tid}")
async def admin_get_tenant(tid: str, user: dict = Depends(get_current_user)):
    _is_super(user)
    owner = await db.users.find_one({"id": tid, "role": "owner"}, {"_id": 0, "password_hash": 0})
    if not owner:
        raise HTTPException(404, "Tenant not found")
    metrics = await _tenant_metrics(tid)
    recent_orders = await db.orders.find({"owner_id": tid}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
    recent_customers = await db.customers.find({"owner_id": tid}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
    branches = await db.branches.find({"owner_id": tid}, {"_id": 0}).to_list(50)
    return {"tenant": owner, "metrics": metrics, "recent_orders": recent_orders, "recent_customers": recent_customers, "branches": branches}


@api.post("/admin/tenants/{tid}/status")
async def admin_set_tenant_status(tid: str, body: TenantStatusIn, user: dict = Depends(get_current_user)):
    _is_super(user)
    owner = await db.users.find_one({"id": tid, "role": "owner"})
    if not owner:
        raise HTTPException(404, "Tenant not found")
    await db.users.update_one({"id": tid}, {"$set": {"status": body.status, "status_changed_at": now_utc()}})
    # Also suspend/restore any sub-users
    await db.users.update_many({"owner_id": tid}, {"$set": {"status": body.status}})
    return {"ok": True, "status": body.status}


@api.post("/admin/tenants/{tid}/subscription")
async def admin_override_subscription(tid: str, body: TenantPlanIn, user: dict = Depends(get_current_user)):
    _is_super(user)
    owner = await db.users.find_one({"id": tid, "role": "owner"})
    if not owner:
        raise HTTPException(404, "Tenant not found")
    plan = SUB_PLANS.get(body.plan_id)
    if not plan:
        raise HTTPException(404, "Unknown plan")
    expires = now_utc() + timedelta(days=int(body.days or 30))
    update = {
        "plan_id": body.plan_id,
        "status": "active",
        "started_at": now_utc(),
        "expires_at": expires,
        "auto_renew": False,
        "granted_by_admin": user.get("email"),
    }
    note = {
        "id": f"admin_grant_{uuid.uuid4().hex[:10]}",
        "by_admin": user.get("email"),
        "plan_id": body.plan_id,
        "days": int(body.days or 30),
        "at": now_utc(),
        "status": "complimentary",
    }
    await db.subscriptions.update_one(
        {"user_id": tid},
        {"$set": update, "$push": {"billing_history": note}, "$setOnInsert": {"id": str(uuid.uuid4()), "user_id": tid}},
        upsert=True,
    )
    sub = await db.subscriptions.find_one({"user_id": tid}, {"_id": 0})
    return {**sub, "plan": plan}


@api.delete("/admin/tenants/{tid}")
async def admin_delete_tenant(tid: str, user: dict = Depends(get_current_user)):
    _is_super(user)
    owner = await db.users.find_one({"id": tid, "role": "owner"})
    if not owner:
        raise HTTPException(404, "Tenant not found")
    # cascade delete tenant data
    for coll in ["customers", "inventory", "orders", "branches", "reminders"]:
        await db[coll].delete_many({"owner_id": tid})
    await db.subscriptions.delete_many({"user_id": tid})
    await db.users.delete_many({"$or": [{"id": tid}, {"owner_id": tid}]})
    return {"ok": True, "deleted_tenant": tid}


@api.get("/admin/recent-signups")
async def admin_recent_signups(limit: int = 20, user: dict = Depends(get_current_user)):
    _is_super(user)
    return await db.users.find({"role": "owner"}, {"_id": 0, "password_hash": 0}).sort("created_at", -1).limit(limit).to_list(limit)


@api.post("/admin/broadcast")
async def admin_broadcast(body: AdminBroadcastIn, user: dict = Depends(get_current_user)):
    _is_super(user)
    rec = {
        "id": str(uuid.uuid4()),
        "title": body.title,
        "message": body.message,
        "severity": body.severity,
        "by": user.get("email"),
        "created_at": now_utc(),
    }
    await db.broadcasts.insert_one(rec.copy())
    return doc(rec)


@api.get("/admin/broadcasts")
async def admin_list_broadcasts(user: dict = Depends(get_current_user)):
    _is_super(user)
    return await db.broadcasts.find({}, {"_id": 0}).sort("created_at", -1).limit(50).to_list(50)


# Public: any logged-in tenant fetches latest broadcast for in-app banner
@api.get("/broadcasts/latest")
async def latest_broadcast(user: dict = Depends(get_current_user)):
    b = await db.broadcasts.find_one({}, {"_id": 0}, sort=[("created_at", -1)])
    return b or {}


# ---------- Inventory CSV Export / Import (admin/owner only) ----------
INVENTORY_CSV_HEADERS = [
    "name", "category", "brand", "model", "color", "shape", "material",
    "lens_index", "blue_cut", "photochromic", "progressive_lens", "coatings",
    "cost", "price", "mrp", "gst_rate", "hsn_code",
    "stock", "low_stock_threshold", "rack_location", "supplier",
    "warranty_months", "sku", "barcode", "branch_id",
]


@api.get("/inventory.csv")
async def inventory_csv_export(user: dict = Depends(require_roles("super_admin", "owner", "admin"))):
    cursor = db.inventory.find(tenant_filter(user), {"_id": 0}).sort("name", 1)
    items = await cursor.to_list(20000)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(INVENTORY_CSV_HEADERS)
    for it in items:
        w.writerow([
            it.get("name", ""), it.get("category", ""), it.get("brand", ""), it.get("model", ""),
            it.get("color", ""), it.get("shape", ""), it.get("material", ""),
            it.get("lens_index") if it.get("lens_index") is not None else "",
            "yes" if it.get("blue_cut") else "no",
            "yes" if it.get("photochromic") else "no",
            "yes" if it.get("progressive_lens") else "no",
            it.get("coatings", ""),
            it.get("cost", 0), it.get("price", 0), it.get("mrp") if it.get("mrp") is not None else "",
            it.get("gst_rate") if it.get("gst_rate") is not None else "",
            it.get("hsn_code", ""),
            it.get("stock", 0), it.get("low_stock_threshold", 3),
            it.get("rack_location", ""), it.get("supplier", ""),
            it.get("warranty_months", 0), it.get("sku", ""), it.get("barcode", ""),
            it.get("branch_id", "") or "",
        ])
    return Response(buf.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=inventory.csv"})


@api.get("/inventory-template.csv")
async def inventory_csv_template(user: dict = Depends(require_roles("super_admin", "owner", "admin"))):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["name", "category", "brand", "model", "color", "price", "cost", "stock", "low_stock_threshold", "sku", "barcode"])
    w.writerow(["Ray-Ban Wayfarer", "frame", "Ray-Ban", "RB2140", "Black", 6500, 3200, 12, 3, "RB-2140-BLK", "8901234567890"])
    w.writerow(["Essilor Crizal 1.6", "lens", "Essilor", "Crizal", "Clear", 2800, 1400, 25, 5, "ESS-CRZ-160", ""])
    return Response(buf.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=inventory_template.csv"})


@api.post("/inventory-import")
async def inventory_import(body: InventoryImportIn, user: dict = Depends(require_roles("super_admin", "owner", "admin"))):
    oid = current_owner_id(user)
    imported, skipped = 0, 0
    errors: List[Dict[str, Any]] = []

    existing_skus = set()
    if body.skip_duplicates:
        cursor = db.inventory.find(tenant_filter(user), {"_id": 0, "sku": 1})
        for it in await cursor.to_list(50000):
            sku = (it.get("sku") or "").strip()
            if sku:
                existing_skus.add(sku.lower())

    to_insert: List[Dict[str, Any]] = []
    for idx, row in enumerate(body.rows):
        try:
            if not row.name or not row.category or row.price is None:
                errors.append({"row": idx + 1, "error": "name, category, and price are required"})
                continue
            sku = (row.sku or "").strip()
            if body.skip_duplicates and sku and sku.lower() in existing_skus:
                skipped += 1
                continue
            if sku:
                existing_skus.add(sku.lower())
            payload = row.model_dump()
            if payload.get("gst_rate") is None:
                payload["gst_rate"] = DEFAULT_GST_RATE
            to_insert.append({
                "id": str(uuid.uuid4()),
                **payload,
                "owner_id": oid,
                "created_at": now_utc(),
            })
            imported += 1
        except Exception as e:
            errors.append({"row": idx + 1, "error": str(e)})

    if to_insert:
        await db.inventory.insert_many([dict(d) for d in to_insert])

    return {"imported": imported, "skipped_duplicates": skipped, "errors": errors, "total_received": len(body.rows)}


# ---------- Sales (historical) CSV Import (admin/owner only) ----------
@api.get("/sales-template.csv")
async def sales_csv_template(user: dict = Depends(require_roles("super_admin", "owner", "admin"))):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["invoice_no", "date", "customer_name", "customer_phone", "subtotal", "gst_amount", "discount", "total", "paid", "payment_status", "notes"])
    w.writerow(["INV-2025-0001", "2025-12-01", "Ravi Kumar", "+919900000001", 5400, 600, 0, 6000, 6000, "paid", "Migrated from spreadsheet"])
    w.writerow(["INV-2025-0002", "2025-12-12", "Priya Sharma", "+919900000002", 3500, 420, 200, 3720, 2000, "partial", ""])
    return Response(buf.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=sales_template.csv"})


@api.post("/sales-import")
async def sales_import(body: SalesImportIn, user: dict = Depends(require_roles("super_admin", "owner", "admin"))):
    oid = current_owner_id(user)
    imported, skipped = 0, 0
    errors: List[Dict[str, Any]] = []

    existing_inv = set()
    if body.skip_duplicates:
        cursor = db.orders.find(tenant_filter(user), {"_id": 0, "invoice_no": 1})
        for o in await cursor.to_list(50000):
            iv = (o.get("invoice_no") or "").strip()
            if iv:
                existing_inv.add(iv.lower())

    to_insert: List[Dict[str, Any]] = []
    for idx, row in enumerate(body.rows):
        try:
            if not row.customer_name or row.total is None:
                errors.append({"row": idx + 1, "error": "customer_name and total are required"})
                continue
            inv = (row.invoice_no or "").strip()
            if body.skip_duplicates and inv and inv.lower() in existing_inv:
                skipped += 1
                continue
            if inv:
                existing_inv.add(inv.lower())

            # Parse date — accept YYYY-MM-DD or ISO; default to now when missing.
            dt = now_utc()
            if row.date:
                parsed_ok = False
                try:
                    parsed = datetime.fromisoformat(row.date.strip().replace("Z", "+00:00"))
                    if parsed.tzinfo is None:
                        parsed = parsed.replace(tzinfo=timezone.utc)
                    dt = parsed
                    parsed_ok = True
                except Exception:
                    try:
                        dt = datetime.strptime(row.date.strip(), "%Y-%m-%d").replace(tzinfo=timezone.utc)
                        parsed_ok = True
                    except Exception:
                        pass
                if not parsed_ok:
                    errors.append({"row": idx + 1, "error": f"invalid date '{row.date}'"})
                    continue

            total = float(row.total)
            paid = float(row.paid or 0)
            due = max(0.0, total - paid)
            status = row.payment_status or ("paid" if due == 0 else ("partial" if paid > 0 else "unpaid"))
            subtotal = float(row.subtotal) if row.subtotal is not None else max(0.0, total - float(row.gst_amount or 0) + float(row.discount or 0))

            to_insert.append({
                "id": str(uuid.uuid4()),
                "invoice_no": inv or f"IMP-{uuid.uuid4().hex[:8].upper()}",
                "customer_id": None,
                "customer_name": row.customer_name,
                "customer_phone": row.customer_phone or "",
                "lines": [],  # historical — no line items
                "subtotal": round(subtotal, 2),
                "gst_amount": float(row.gst_amount or 0),
                "discount": float(row.discount or 0),
                "total": round(total, 2),
                "paid": round(paid, 2),
                "due": round(due, 2),
                "payment_status": status,
                "fulfillment_status": "delivered",
                "notes": row.notes or "",
                "is_imported": True,
                "branch_id": row.branch_id,
                "owner_id": oid,
                "created_at": dt,
            })
            imported += 1
        except Exception as e:
            errors.append({"row": idx + 1, "error": str(e)})

    if to_insert:
        await db.orders.insert_many([dict(d) for d in to_insert])

    return {"imported": imported, "skipped_duplicates": skipped, "errors": errors, "total_received": len(body.rows)}


# ---------- Customer CSV Export / Import (admin/owner only) ----------
CUSTOMER_CSV_HEADERS = ["name", "phone", "email", "address", "dob", "gstin", "notes", "branch_id", "loyalty_points", "created_at"]


@api.get("/customers.csv")
async def customers_csv_export(user: dict = Depends(require_roles("super_admin", "owner", "admin"))):
    cursor = db.customers.find(tenant_filter(user), {"_id": 0}).sort("created_at", -1)
    customers = await cursor.to_list(10000)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(CUSTOMER_CSV_HEADERS)
    for c in customers:
        w.writerow([
            c.get("name", ""),
            c.get("phone", ""),
            c.get("email", ""),
            c.get("address", ""),
            c.get("dob", ""),
            c.get("gstin", ""),
            (c.get("notes", "") or "").replace("\n", " "),
            c.get("branch_id", "") or "",
            c.get("loyalty_points", 0),
            c["created_at"].strftime("%Y-%m-%d %H:%M") if isinstance(c.get("created_at"), datetime) else c.get("created_at", ""),
        ])
    return Response(
        buf.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=customers.csv"},
    )


@api.get("/customers-template.csv")
async def customers_csv_template(user: dict = Depends(require_roles("super_admin", "owner", "admin"))):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["name", "phone", "email", "address", "dob", "gstin", "notes"])
    w.writerow(["John Doe", "+919999999999", "john@example.com", "MG Road", "1990-01-15", "", "VIP customer"])
    return Response(
        buf.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=customers_template.csv"},
    )


@api.post("/customers-import")
async def customers_import(body: CustomerImportIn, user: dict = Depends(require_roles("super_admin", "owner", "admin"))):
    oid = current_owner_id(user)
    imported = 0
    skipped = 0
    errors: List[Dict[str, Any]] = []

    # Pre-fetch existing phones for duplicate detection
    existing = set()
    if body.skip_duplicates:
        cursor = db.customers.find(tenant_filter(user), {"_id": 0, "phone": 1})
        for c in await cursor.to_list(20000):
            if c.get("phone"):
                existing.add(c["phone"].strip())

    to_insert: List[Dict[str, Any]] = []
    for idx, row in enumerate(body.rows):
        try:
            phone = (row.phone or "").strip()
            name = (row.name or "").strip()
            if not name or not phone:
                errors.append({"row": idx + 1, "error": "name and phone are required"})
                continue
            if body.skip_duplicates and phone in existing:
                skipped += 1
                continue
            existing.add(phone)
            to_insert.append({
                "id": str(uuid.uuid4()),
                "name": name,
                "phone": phone,
                "email": (row.email or "").strip(),
                "address": row.address or "",
                "dob": row.dob or "",
                "birthday": row.birthday or "",
                "anniversary": row.anniversary or "",
                "gstin": row.gstin or "",
                "notes": row.notes or "",
                "branch_id": row.branch_id,
                "owner_id": oid,
                "prescriptions": [],
                "loyalty_points": 0,
                "created_at": now_utc(),
                "last_visit": now_utc(),
            })
            imported += 1
        except Exception as e:
            errors.append({"row": idx + 1, "error": str(e)})

    if to_insert:
        await db.customers.insert_many([dict(d) for d in to_insert])

    return {"imported": imported, "skipped_duplicates": skipped, "errors": errors, "total_received": len(body.rows)}


# ---------- Staff Management (owner/admin manages their tenant's users) ----------
@api.get("/staff")
async def list_staff(user: dict = Depends(require_roles("super_admin", "owner", "admin"))):
    oid = current_owner_id(user)
    query: Dict[str, Any] = {"role": {"$in": ["admin", "staff", "owner"]}}
    if oid is not None:
        query["owner_id"] = oid
    cursor = db.users.find(query, {"_id": 0, "password_hash": 0}).sort("created_at", -1).limit(500)
    return await cursor.to_list(500)


@api.post("/staff")
async def create_staff(body: StaffCreateIn, user: dict = Depends(require_roles("super_admin", "owner", "admin"))):
    existing = await db.users.find_one({"email": body.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    oid = current_owner_id(user) or user["id"]
    uid = str(uuid.uuid4())
    new_user = {
        "id": uid,
        "email": body.email.lower(),
        "name": body.name,
        "password_hash": hash_password(body.password),
        "role": body.role,
        "owner_id": oid,
        "status": "active",
        "branch_id": body.branch_id,
        "created_at": now_utc(),
        "created_by": user.get("email"),
    }
    await db.users.insert_one(new_user.copy())
    return {
        "id": uid,
        "email": new_user["email"],
        "name": new_user["name"],
        "role": new_user["role"],
        "owner_id": oid,
        "branch_id": body.branch_id,
        "status": "active",
        "created_at": new_user["created_at"],
    }


@api.put("/staff/{sid}")
async def update_staff(sid: str, body: StaffUpdateIn, user: dict = Depends(require_roles("super_admin", "owner", "admin"))):
    oid = current_owner_id(user)
    query: Dict[str, Any] = {"id": sid}
    if oid is not None:
        query["owner_id"] = oid
    target = await db.users.find_one(query)
    if not target:
        raise HTTPException(404, "Staff not found")
    if target.get("role") == "owner" and target["id"] != user["id"]:
        raise HTTPException(403, "Cannot modify owner")
    # Only apply fields the client explicitly sent (so branch_id: null can clear assignment).
    update: Dict[str, Any] = body.model_dump(exclude_unset=True)
    if "password" in update:
        pwd = update.pop("password")
        if pwd:
            update["password_hash"] = hash_password(pwd)
    if update:
        await db.users.update_one({"id": sid}, {"$set": update})
    out = await db.users.find_one({"id": sid}, {"_id": 0, "password_hash": 0})
    return out


@api.delete("/staff/{sid}")
async def delete_staff(sid: str, user: dict = Depends(require_roles("super_admin", "owner", "admin"))):
    if sid == user["id"]:
        raise HTTPException(400, "Cannot delete yourself")
    oid = current_owner_id(user)
    query: Dict[str, Any] = {"id": sid}
    if oid is not None:
        query["owner_id"] = oid
    target = await db.users.find_one(query)
    if not target:
        raise HTTPException(404, "Staff not found")
    if target.get("role") == "owner":
        raise HTTPException(403, "Cannot delete owner")
    await db.users.delete_one({"id": sid})
    return {"ok": True}


# ---------- Sync (prefetch all core data after login) ----------
@api.get("/sync")
async def sync_all(branch_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    """One-shot prefetch: dashboard, customers, inventory, orders, branches, reminders, broadcasts.
    Used by clients to warm the offline cache on login / app resume."""
    base: Dict[str, Any] = tenant_filter(user)
    base_branch: Dict[str, Any] = dict(base)
    if branch_id:
        base_branch["branch_id"] = branch_id

    today = now_utc().replace(hour=0, minute=0, second=0, microsecond=0)
    month_start = today.replace(day=1)

    branches = await db.branches.find(base, {"_id": 0}).sort("name", 1).to_list(500)
    customers = await db.customers.find(base_branch, {"_id": 0}).sort("created_at", -1).limit(2000).to_list(2000)
    inventory = await db.inventory.find(base_branch, {"_id": 0}).sort("name", 1).limit(5000).to_list(5000)
    orders = await db.orders.find(base_branch, {"_id": 0}).sort("created_at", -1).limit(1000).to_list(1000)
    reminders = await db.reminders.find(base, {"_id": 0}).sort("sent_at", -1).limit(200).to_list(200)
    broadcast = await db.broadcasts.find_one({}, {"_id": 0}, sort=[("created_at", -1)])

    orders_today = [o for o in orders if isinstance(o.get("created_at"), datetime) and o["created_at"] >= today]
    orders_month = [o for o in orders if isinstance(o.get("created_at"), datetime) and o["created_at"] >= month_start]
    open_orders = [o for o in orders if o.get("payment_status") != "paid"]
    low_stock = [i for i in inventory if i.get("stock", 0) <= i.get("low_stock_threshold", 3)][:10]

    pipeline = {}
    for s in ORDER_STATUSES:
        pipeline[s] = sum(1 for o in orders if o.get("fulfillment_status") == s)

    # staff list (admin-only side info — silently empty for staff role)
    staff: List[Dict[str, Any]] = []
    if user.get("role") in ("super_admin", "owner", "admin"):
        sq: Dict[str, Any] = {"role": {"$in": ["admin", "staff", "owner"]}}
        oid = current_owner_id(user)
        if oid is not None:
            sq["owner_id"] = oid
        staff = await db.users.find(sq, {"_id": 0, "password_hash": 0}).sort("created_at", -1).limit(200).to_list(200)

    return {
        "synced_at": now_utc(),
        "branches": branches,
        "customers": customers,
        "inventory": inventory,
        "orders": orders,
        "reminders": reminders,
        "broadcast": broadcast or {},
        "staff": staff,
        "dashboard": {
            "revenue_today": sum(o.get("paid", 0) for o in orders_today),
            "revenue_month": sum(o.get("paid", 0) for o in orders_month),
            "pending_due": sum(o.get("due", 0) for o in open_orders),
            "gst_collected_month": round(sum(o.get("gst_amount", 0) for o in orders_month), 2),
            "orders_today": len(orders_today),
            "orders_month": len(orders_month),
            "customers_count": len(customers),
            "inventory_count": len(inventory),
            "low_stock": low_stock,
            "recent_customers": customers[:5],
            "recent_orders": orders[:5],
            "pipeline": pipeline,
        },
    }


# ---------- Repair Orders ----------
class RepairOrderIn(BaseModel):
    customer_id: str
    item_description: str
    issue: str
    estimated_cost: float = 0
    advance_paid: float = 0
    expected_date: Optional[str] = ""
    notes: Optional[str] = ""
    branch_id: Optional[str] = None


class RepairStatusIn(BaseModel):
    status: Literal["received", "diagnosed", "in_repair", "ready", "delivered", "cancelled"]
    note: Optional[str] = ""


@api.get("/repair-orders")
async def list_repair_orders(branch_id: Optional[str] = None, status: Optional[str] = None, q: Optional[str] = None, user: dict = Depends(get_current_user)):
    query: Dict[str, Any] = tenant_filter(user)
    if branch_id:
        query["branch_id"] = branch_id
    if status:
        query["status"] = status
    if q:
        query["$or"] = [
            {"customer_name": {"$regex": q, "$options": "i"}},
            {"item_description": {"$regex": q, "$options": "i"}},
            {"repair_no": {"$regex": q, "$options": "i"}},
        ]
    cursor = db.repair_orders.find(query, {"_id": 0}).sort("created_at", -1).limit(1000)
    return await cursor.to_list(1000)


@api.post("/repair-orders")
async def create_repair_order(body: RepairOrderIn, user: dict = Depends(get_current_user)):
    customer = await db.customers.find_one(tenant_filter(user, {"id": body.customer_id}), {"_id": 0})
    if not customer:
        raise HTTPException(404, "Customer not found")
    counter = await db.counters.find_one_and_update(
        {"id": "repair"}, {"$inc": {"seq": 1}}, upsert=True, return_document=True,
    )
    seq = (counter or {}).get("seq", 1)
    rec = {
        "id": str(uuid.uuid4()),
        "repair_no": f"RPR-{now_utc().strftime('%Y%m')}-{seq:05d}",
        "owner_id": current_owner_id(user),
        "customer_id": body.customer_id,
        "customer_name": customer.get("name"),
        "customer_phone": customer.get("phone"),
        "item_description": body.item_description,
        "issue": body.issue,
        "estimated_cost": float(body.estimated_cost or 0),
        "advance_paid": float(body.advance_paid or 0),
        "final_cost": 0.0,
        "expected_date": body.expected_date or "",
        "notes": body.notes or "",
        "branch_id": body.branch_id,
        "status": "received",
        "timeline": [{"status": "received", "at": now_utc(), "note": "Repair received", "by": user.get("name")}],
        "created_at": now_utc(),
    }
    await db.repair_orders.insert_one(rec.copy())
    return doc(rec)


@api.get("/repair-orders/{rid}")
async def get_repair_order(rid: str, user: dict = Depends(get_current_user)):
    r = await db.repair_orders.find_one(tenant_filter(user, {"id": rid}), {"_id": 0})
    if not r:
        raise HTTPException(404, "Repair order not found")
    return r


@api.post("/repair-orders/{rid}/status")
async def update_repair_status(rid: str, body: RepairStatusIn, user: dict = Depends(get_current_user)):
    r = await db.repair_orders.find_one(tenant_filter(user, {"id": rid}))
    if not r:
        raise HTTPException(404, "Repair order not found")
    event = {"status": body.status, "at": now_utc(), "note": body.note or "", "by": user.get("name")}
    await db.repair_orders.update_one({"id": rid}, {"$set": {"status": body.status}, "$push": {"timeline": event}})
    return await db.repair_orders.find_one({"id": rid}, {"_id": 0})


@api.delete("/repair-orders/{rid}")
async def delete_repair_order(rid: str, user: dict = Depends(get_current_user)):
    await db.repair_orders.delete_one(tenant_filter(user, {"id": rid}))
    return {"ok": True}


# ---------- Coupon Codes ----------
class CouponIn(BaseModel):
    code: str
    discount_type: Literal["percent", "flat"] = "percent"
    value: float
    min_order: float = 0
    max_discount: Optional[float] = None
    expires_at: Optional[str] = ""  # ISO date
    usage_limit: int = 0  # 0 = unlimited
    active: bool = True
    description: Optional[str] = ""


@api.get("/coupons")
async def list_coupons(user: dict = Depends(get_current_user)):
    return await db.coupons.find(tenant_filter(user), {"_id": 0}).sort("created_at", -1).to_list(500)


@api.post("/coupons")
async def create_coupon(body: CouponIn, user: dict = Depends(require_roles("super_admin", "owner", "admin"))):
    code = body.code.strip().upper()
    if not code:
        raise HTTPException(400, "Code is required")
    existing = await db.coupons.find_one(tenant_filter(user, {"code": code}))
    if existing:
        raise HTTPException(400, "Coupon code already exists")
    rec = {
        "id": str(uuid.uuid4()),
        "owner_id": current_owner_id(user),
        **body.model_dump(),
        "code": code,
        "uses": 0,
        "created_at": now_utc(),
        "created_by": user.get("email"),
    }
    await db.coupons.insert_one(rec.copy())
    return doc(rec)


@api.put("/coupons/{cid}")
async def update_coupon(cid: str, body: CouponIn, user: dict = Depends(require_roles("super_admin", "owner", "admin"))):
    update = body.model_dump()
    update["code"] = update["code"].strip().upper()
    res = await db.coupons.update_one(tenant_filter(user, {"id": cid}), {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(404, "Coupon not found")
    return await db.coupons.find_one({"id": cid}, {"_id": 0})


@api.delete("/coupons/{cid}")
async def delete_coupon(cid: str, user: dict = Depends(require_roles("super_admin", "owner", "admin"))):
    await db.coupons.delete_one(tenant_filter(user, {"id": cid}))
    return {"ok": True}


@api.post("/coupons/validate")
async def validate_coupon(code: str, subtotal: float, user: dict = Depends(get_current_user)):
    c = await db.coupons.find_one(tenant_filter(user, {"code": code.strip().upper()}), {"_id": 0})
    if not c:
        raise HTTPException(404, "Coupon not found")
    if not c.get("active"):
        raise HTTPException(400, "Coupon inactive")
    if c.get("expires_at"):
        try:
            exp = datetime.fromisoformat(c["expires_at"]).replace(tzinfo=timezone.utc) if c["expires_at"] else None
            if exp and exp < now_utc():
                raise HTTPException(400, "Coupon expired")
        except ValueError:
            pass
    if c.get("usage_limit") and c.get("uses", 0) >= c["usage_limit"]:
        raise HTTPException(400, "Coupon usage limit reached")
    if subtotal < float(c.get("min_order", 0)):
        raise HTTPException(400, f"Minimum order amount is {c.get('min_order')}")
    if c["discount_type"] == "percent":
        disc = round(subtotal * float(c["value"]) / 100.0, 2)
        if c.get("max_discount"):
            disc = min(disc, float(c["max_discount"]))
    else:
        disc = float(c["value"])
    return {"valid": True, "discount": disc, "coupon": c}


# ---------- Referral System ----------
class ReferralIn(BaseModel):
    referrer_customer_id: str
    referred_name: str
    referred_phone: str
    notes: Optional[str] = ""


@api.get("/referrals")
async def list_referrals(user: dict = Depends(get_current_user)):
    return await db.referrals.find(tenant_filter(user), {"_id": 0}).sort("created_at", -1).to_list(500)


@api.post("/referrals")
async def create_referral(body: ReferralIn, user: dict = Depends(get_current_user)):
    referrer = await db.customers.find_one(tenant_filter(user, {"id": body.referrer_customer_id}), {"_id": 0})
    if not referrer:
        raise HTTPException(404, "Referring customer not found")
    rec = {
        "id": str(uuid.uuid4()),
        "owner_id": current_owner_id(user),
        "referrer_customer_id": body.referrer_customer_id,
        "referrer_name": referrer.get("name"),
        "referrer_phone": referrer.get("phone"),
        "referred_name": body.referred_name,
        "referred_phone": body.referred_phone,
        "status": "pending",  # pending | converted | rewarded
        "reward_points": 0,
        "notes": body.notes or "",
        "created_at": now_utc(),
    }
    await db.referrals.insert_one(rec.copy())
    return doc(rec)


@api.post("/referrals/{rid}/convert")
async def convert_referral(rid: str, reward_points: int = 100, user: dict = Depends(get_current_user)):
    r = await db.referrals.find_one(tenant_filter(user, {"id": rid}))
    if not r:
        raise HTTPException(404, "Referral not found")
    await db.referrals.update_one({"id": rid}, {"$set": {"status": "rewarded", "reward_points": reward_points, "converted_at": now_utc()}})
    # Credit referrer with loyalty points
    await db.customers.update_one({"id": r["referrer_customer_id"]}, {"$inc": {"loyalty_points": reward_points}})
    return await db.referrals.find_one({"id": rid}, {"_id": 0})


@api.delete("/referrals/{rid}")
async def delete_referral(rid: str, user: dict = Depends(get_current_user)):
    await db.referrals.delete_one(tenant_filter(user, {"id": rid}))
    return {"ok": True}


# ---------- Subscription: Auto-Renewal + Expiry Reminder ----------
class AutoRenewIn(BaseModel):
    auto_renew: bool
    reminder_days: Optional[int] = 7  # send expiry reminder N days before


@api.post("/subscription/auto-renew")
async def set_auto_renew(body: AutoRenewIn, user: dict = Depends(get_current_user)):
    await db.subscriptions.update_one(
        {"user_id": user["id"]},
        {"$set": {"auto_renew": body.auto_renew, "reminder_days": int(body.reminder_days or 7)}},
        upsert=True,
    )
    sub = await db.subscriptions.find_one({"user_id": user["id"]}, {"_id": 0})
    return sub


@api.get("/subscription/expiry-reminder")
async def expiry_reminder(user: dict = Depends(get_current_user)):
    """Check if my subscription is expiring within the reminder window."""
    sub = await db.subscriptions.find_one({"user_id": user["id"]}, {"_id": 0})
    if not sub or not sub.get("expires_at"):
        return {"expiring_soon": False}
    exp = sub["expires_at"] if isinstance(sub["expires_at"], datetime) else datetime.fromisoformat(str(sub["expires_at"]))
    if exp.tzinfo is None:
        exp = exp.replace(tzinfo=timezone.utc)
    days_left = (exp - now_utc()).days
    reminder_days = int(sub.get("reminder_days") or 7)
    return {
        "expiring_soon": days_left <= reminder_days,
        "days_left": days_left,
        "expires_at": exp.isoformat(),
        "plan_id": sub.get("plan_id"),
        "auto_renew": sub.get("auto_renew", False),
    }


# ---------- Birthday / Anniversary Wishes ----------
def _md_key(s: str) -> Optional[str]:
    """Return MM-DD from a YYYY-MM-DD string."""
    if not s or len(s) < 5:
        return None
    parts = s.split("-")
    if len(parts) >= 3:
        return f"{parts[1]}-{parts[2][:2]}"
    return None


@api.get("/customers/celebrations/today")
async def celebrations_today(user: dict = Depends(get_current_user)):
    today_md = now_utc().strftime("%m-%d")
    customers = await db.customers.find(tenant_filter(user), {"_id": 0}).to_list(20000)
    birthdays = [c for c in customers if _md_key(c.get("birthday") or "") == today_md or _md_key(c.get("dob") or "") == today_md]
    anniversaries = [c for c in customers if _md_key(c.get("anniversary") or "") == today_md]
    return {
        "date": now_utc().strftime("%Y-%m-%d"),
        "birthdays": birthdays,
        "anniversaries": anniversaries,
    }


class WishSendIn(BaseModel):
    customer_id: str
    channel: Literal["sms", "whatsapp", "email"] = "whatsapp"
    occasion: Literal["birthday", "anniversary"]
    message: Optional[str] = ""


@api.post("/customers/wishes/send")
async def send_wish(body: WishSendIn, user: dict = Depends(get_current_user)):
    c = await db.customers.find_one(tenant_filter(user, {"id": body.customer_id}), {"_id": 0})
    if not c:
        raise HTTPException(404, "Customer not found")
    if not body.message:
        if body.occasion == "birthday":
            msg = f"Dear {c.get('name')}, wishing you a very Happy Birthday from ARN Optical! Visit us this week for a special birthday discount."
        else:
            msg = f"Dear {c.get('name')}, Happy Anniversary! Wishing you many more years of happiness from ARN Optical."
    else:
        msg = body.message
    rec = {
        "id": str(uuid.uuid4()),
        "owner_id": current_owner_id(user),
        "customer_id": body.customer_id,
        "customer_name": c.get("name"),
        "channel": body.channel,
        "message": msg,
        "phone": c.get("phone"),
        "occasion": body.occasion,
        "kind": "wish",
        "status": "sent_mock",
        "sent_at": now_utc(),
    }
    await db.reminders.insert_one(rec.copy())
    return doc(rec)


@api.post("/customers/wishes/send-bulk")
async def send_wishes_bulk(occasion: Literal["birthday", "anniversary"], channel: Literal["sms", "whatsapp", "email"] = "whatsapp", user: dict = Depends(get_current_user)):
    today_md = now_utc().strftime("%m-%d")
    customers = await db.customers.find(tenant_filter(user), {"_id": 0}).to_list(20000)
    sent = 0
    if occasion == "birthday":
        targets = [c for c in customers if _md_key(c.get("birthday") or "") == today_md or _md_key(c.get("dob") or "") == today_md]
    else:
        targets = [c for c in customers if _md_key(c.get("anniversary") or "") == today_md]
    for c in targets:
        if occasion == "birthday":
            msg = f"Dear {c.get('name')}, wishing you a very Happy Birthday from ARN Optical!"
        else:
            msg = f"Dear {c.get('name')}, Happy Anniversary from ARN Optical!"
        rec = {
            "id": str(uuid.uuid4()),
            "owner_id": current_owner_id(user),
            "customer_id": c.get("id"),
            "customer_name": c.get("name"),
            "channel": channel,
            "message": msg,
            "phone": c.get("phone"),
            "occasion": occasion,
            "kind": "wish",
            "status": "sent_mock",
            "sent_at": now_utc(),
        }
        await db.reminders.insert_one(rec)
        sent += 1
    return {"sent": sent, "occasion": occasion, "channel": channel}


# ---------- AI Prescription Scanner ----------
class PrescriptionScanIn(BaseModel):
    image_base64: str  # base64 of the prescription image (without data: prefix)
    mime_type: Optional[str] = "image/jpeg"


@api.post("/prescription/ai-scan")
async def ai_scan_prescription(body: PrescriptionScanIn, user: dict = Depends(get_current_user)):
    """Vision-based prescription scanner — extracts OD/OS/PD values from prescription paper image."""
    if not EMERGENT_LLM_KEY:
        raise HTTPException(503, "AI key not configured")
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"rx-scan-{uuid.uuid4().hex[:8]}",
            system_message=(
                "You are an optometry data-extraction assistant. The user uploads a photo of a written or printed eye prescription. "
                "Extract numerical values into strict JSON with this schema:\n"
                "{\n"
                "  \"od_sph\": number|null, \"od_cyl\": number|null, \"od_axis\": number|null, \"od_add\": number|null, \"od_va\": string|null,\n"
                "  \"os_sph\": number|null, \"os_cyl\": number|null, \"os_axis\": number|null, \"os_add\": number|null, \"os_va\": string|null,\n"
                "  \"pd\": number|null, \"near_pd\": number|null,\n"
                "  \"rx_type\": \"distance\"|\"near\"|\"intermediate\"|\"progressive\"|\"bifocal\"|\"contact\"|\"pediatric\",\n"
                "  \"doctor_name\": string|null, \"diagnosis\": string|null, \"date\": string|null,\n"
                "  \"confidence\": number (0-1),\n"
                "  \"notes\": string\n"
                "}\n"
                "Return ONLY JSON. No prose. If a field is unreadable, set it to null. Numbers must be parsed (e.g., -1.25, not '-1.25')."
            ),
        ).with_model("gemini", "gemini-3-flash-preview")
        img = ImageContent(image_base64=body.image_base64)
        resp = await chat.send_message(UserMessage(text="Extract the prescription as JSON.", file_contents=[img]))
        text = str(resp).strip()
        # Try to clean possible code-fences
        import json as _json
        if text.startswith("```"):
            text = text.strip("`")
            if text.lower().startswith("json"):
                text = text[4:].strip()
        try:
            parsed = _json.loads(text)
        except Exception:
            # Find first JSON object in text
            start = text.find("{")
            end = text.rfind("}")
            if start >= 0 and end > start:
                parsed = _json.loads(text[start:end + 1])
            else:
                raise
        return {"ok": True, "extracted": parsed, "raw": text[:2000]}
    except Exception as e:
        logging.exception("AI scan failed")
        err = str(e)
        friendly = "AI service unavailable. Top up your Emergent LLM key balance." if "budget" in err.lower() else f"AI scan failed: {err[:120]}"
        raise HTTPException(502, friendly)


# ---------- Excel & PDF Exports ----------
def _xlsx_bytes(headers: List[str], rows: List[List[Any]], sheet_name: str = "Sheet1") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:30]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
    for r in rows:
        ws.append(r)
    for col_idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "AA"].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pdf_bytes(title: str, headers: List[str], rows: List[List[Any]]) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors as rl_colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc_pdf = SimpleDocTemplate(buf, pagesize=landscape(A4), title=title, leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story: List[Any] = [Paragraph(f"<b>{title}</b>", styles["Title"]), Spacer(1, 8)]
    data = [headers] + [[str(c) if c is not None else "" for c in r] for r in rows]
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.whitesmoke, rl_colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.25, rl_colors.HexColor("#CBD5E1")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(tbl)
    doc_pdf.build(story)
    return buf.getvalue()


@api.get("/customers.xlsx")
async def customers_xlsx_export(user: dict = Depends(require_roles("super_admin", "owner", "admin"))):
    cs = await db.customers.find(tenant_filter(user), {"_id": 0}).sort("created_at", -1).to_list(20000)
    headers = ["Name", "Phone", "Email", "Address", "DOB", "Birthday", "Anniversary", "GSTIN", "Loyalty Points", "Notes", "Created"]
    rows = [[
        c.get("name", ""), c.get("phone", ""), c.get("email", ""), c.get("address", ""),
        c.get("dob", ""), c.get("birthday", ""), c.get("anniversary", ""),
        c.get("gstin", ""), c.get("loyalty_points", 0), (c.get("notes", "") or "").replace("\n", " "),
        c["created_at"].strftime("%Y-%m-%d %H:%M") if isinstance(c.get("created_at"), datetime) else c.get("created_at", ""),
    ] for c in cs]
    data = _xlsx_bytes(headers, rows, "Customers")
    return Response(data, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=customers.xlsx"})


@api.get("/customers.pdf")
async def customers_pdf_export(user: dict = Depends(require_roles("super_admin", "owner", "admin"))):
    cs = await db.customers.find(tenant_filter(user), {"_id": 0}).sort("created_at", -1).to_list(20000)
    headers = ["Name", "Phone", "Email", "DOB", "Birthday", "Anniversary", "GSTIN", "Points"]
    rows = [[c.get("name", ""), c.get("phone", ""), c.get("email", ""), c.get("dob", ""), c.get("birthday", ""), c.get("anniversary", ""), c.get("gstin", ""), c.get("loyalty_points", 0)] for c in cs]
    data = _pdf_bytes("Customers", headers, rows)
    return Response(data, media_type="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=customers.pdf"})


@api.get("/inventory.xlsx")
async def inventory_xlsx_export(user: dict = Depends(require_roles("super_admin", "owner", "admin"))):
    items = await db.inventory.find(tenant_filter(user), {"_id": 0}).sort("name", 1).to_list(20000)
    headers = ["Name", "Category", "Brand", "Model", "Color", "Price", "Cost", "MRP", "Stock", "Low Threshold", "GST%", "HSN", "SKU", "Barcode"]
    rows = [[
        it.get("name", ""), it.get("category", ""), it.get("brand", ""), it.get("model", ""), it.get("color", ""),
        it.get("price", 0), it.get("cost", 0), it.get("mrp") or "", it.get("stock", 0), it.get("low_stock_threshold", 3),
        it.get("gst_rate", DEFAULT_GST_RATE), it.get("hsn_code", "9004"), it.get("sku", ""), it.get("barcode", ""),
    ] for it in items]
    data = _xlsx_bytes(headers, rows, "Inventory")
    return Response(data, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=inventory.xlsx"})


@api.get("/inventory.pdf")
async def inventory_pdf_export(user: dict = Depends(require_roles("super_admin", "owner", "admin"))):
    items = await db.inventory.find(tenant_filter(user), {"_id": 0}).sort("name", 1).to_list(20000)
    headers = ["Name", "Category", "Brand", "Price", "Cost", "Stock", "GST%", "SKU"]
    rows = [[it.get("name", ""), it.get("category", ""), it.get("brand", ""), it.get("price", 0), it.get("cost", 0), it.get("stock", 0), it.get("gst_rate", DEFAULT_GST_RATE), it.get("sku", "")] for it in items]
    data = _pdf_bytes("Inventory", headers, rows)
    return Response(data, media_type="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=inventory.pdf"})


@api.get("/reports/sales.xlsx")
async def sales_xlsx_export(start: Optional[str] = None, end: Optional[str] = None, branch_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    q: Dict[str, Any] = tenant_filter(user)
    if branch_id:
        q["branch_id"] = branch_id
    if start or end:
        rng = {}
        if start:
            rng["$gte"] = datetime.fromisoformat(start).replace(tzinfo=timezone.utc)
        if end:
            rng["$lte"] = datetime.fromisoformat(end).replace(tzinfo=timezone.utc)
        q["created_at"] = rng
    orders = await db.orders.find(q, {"_id": 0}).sort("created_at", -1).to_list(20000)
    headers = ["Invoice", "Date", "Customer", "Phone", "Subtotal", "GST", "Discount", "Total", "Paid", "Due", "Payment", "Fulfillment"]
    rows = [[
        o.get("invoice_no", o["id"][:8]),
        o["created_at"].strftime("%Y-%m-%d %H:%M") if isinstance(o.get("created_at"), datetime) else o.get("created_at"),
        o.get("customer_name"), o.get("customer_phone"),
        o.get("subtotal", 0), o.get("gst_amount", 0), o.get("discount", 0),
        o.get("total", 0), o.get("paid", 0), o.get("due", 0),
        o.get("payment_status"), o.get("fulfillment_status"),
    ] for o in orders]
    data = _xlsx_bytes(headers, rows, "Sales")
    return Response(data, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=sales.xlsx"})


@api.get("/reports/sales.pdf")
async def sales_pdf_export(start: Optional[str] = None, end: Optional[str] = None, branch_id: Optional[str] = None, user: dict = Depends(get_current_user)):
    q: Dict[str, Any] = tenant_filter(user)
    if branch_id:
        q["branch_id"] = branch_id
    if start or end:
        rng = {}
        if start:
            rng["$gte"] = datetime.fromisoformat(start).replace(tzinfo=timezone.utc)
        if end:
            rng["$lte"] = datetime.fromisoformat(end).replace(tzinfo=timezone.utc)
        q["created_at"] = rng
    orders = await db.orders.find(q, {"_id": 0}).sort("created_at", -1).to_list(20000)
    headers = ["Invoice", "Date", "Customer", "Total", "Paid", "Due", "Payment"]
    rows = [[
        o.get("invoice_no", o["id"][:8]),
        o["created_at"].strftime("%Y-%m-%d") if isinstance(o.get("created_at"), datetime) else o.get("created_at"),
        o.get("customer_name"), o.get("total", 0), o.get("paid", 0), o.get("due", 0), o.get("payment_status"),
    ] for o in orders]
    data = _pdf_bytes("Sales Report", headers, rows)
    return Response(data, media_type="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=sales.pdf"})


# ---------- Super-Admin Extensions: Coupons, Referrals, Repair Orders, Wishes ----------
@api.get("/admin/coupons-all")
async def admin_list_all_coupons(user: dict = Depends(get_current_user)):
    _is_super(user)
    return await db.coupons.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)


@api.get("/admin/referrals-all")
async def admin_list_all_referrals(user: dict = Depends(get_current_user)):
    _is_super(user)
    return await db.referrals.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)


@api.get("/admin/repair-orders-all")
async def admin_list_all_repair_orders(user: dict = Depends(get_current_user)):
    _is_super(user)
    return await db.repair_orders.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)


@api.get("/admin/wishes-all")
async def admin_list_all_wishes(user: dict = Depends(get_current_user)):
    _is_super(user)
    return await db.reminders.find({"kind": "wish"}, {"_id": 0}).sort("sent_at", -1).to_list(500)


# ---------- Health ----------
@api.get("/")
async def root():
    return {"name": "OptiCRM API", "status": "ok", "version": "2.0"}


# Mount router and middleware
app.include_router(api)
app.add_middleware(CORSMiddleware, allow_credentials=True, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)


# ---------- Startup: seed ----------
@app.on_event("startup")
async def startup():
    existing = await db.users.find_one({"email": ADMIN_EMAIL.lower()})
    if not existing:
        uid = str(uuid.uuid4())
        await db.users.insert_one({
            "id": uid,
            "email": ADMIN_EMAIL.lower(),
            "name": "Platform Admin",
            "password_hash": hash_password(ADMIN_PASSWORD),
            "role": "super_admin",
            "owner_id": uid,
            "status": "active",
            "branch_id": None,
            "created_at": now_utc(),
        })
        logger.info(f"Seeded super-admin user: {ADMIN_EMAIL}")
    else:
        # ensure admin is super_admin (upgrade from MVP)
        if existing.get("role") != "super_admin":
            await db.users.update_one({"id": existing["id"]}, {"$set": {"role": "super_admin", "status": "active"}})
            logger.info("Upgraded existing admin to super_admin")

    # Seed default branches
    if await db.branches.count_documents({}) == 0:
        await db.branches.insert_many([
            {"id": str(uuid.uuid4()), "name": "ARN Optical — Main Branch", "code": "MAIN", "address": "MG Road, Bengaluru", "phone": "+91 80 0000 0000", "gstin": "29ABCDE1234F1Z5", "created_at": now_utc()},
            {"id": str(uuid.uuid4()), "name": "ARN Optical — Indiranagar", "code": "IND", "address": "Indiranagar, Bengaluru", "phone": "+91 80 1111 1111", "gstin": "29ABCDE1234F1Z5", "created_at": now_utc()},
        ])
        logger.info("Seeded branches")

    # Backfill existing inventory with new fields
    async for it in db.inventory.find({"hsn_code": {"$exists": False}}):
        await db.inventory.update_one({"id": it["id"]}, {"$set": {
            "hsn_code": "9004", "gst_rate": DEFAULT_GST_RATE, "barcode": it.get("sku", ""),
            "rack_location": "", "supplier": "", "warranty_months": 12,
        }})

    # Backfill orders to have fulfillment_status / invoice_no
    async for o in db.orders.find({"fulfillment_status": {"$exists": False}}):
        counter = await db.counters.find_one_and_update({"id": "invoice"}, {"$inc": {"seq": 1}}, upsert=True, return_document=True)
        seq = (counter or {}).get("seq", 1)
        await db.orders.update_one({"id": o["id"]}, {"$set": {
            "fulfillment_status": "delivered" if o.get("status") == "paid" else "received",
            "invoice_no": o.get("invoice_no") or next_invoice_no(seq),
            "gst_amount": 0,
            "timeline": [{"status": "received", "at": o.get("created_at", now_utc()), "note": "Order placed", "by": "system"}],
        }, "$rename": {"status": "payment_status"}})


@app.on_event("shutdown")
async def shutdown():
    client.close()
